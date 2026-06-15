from __future__ import annotations

import argparse
import csv
import gc
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_TEMPLATE_NAME = "Jarir - 3 Statement - Assignment.Sol.xlsx"
REPORT_SHEET_NAME = "Correction Report"
DIFF_FILL = PatternFill(fill_type="solid", fgColor="F4CCCC")
MISSING_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")


@dataclass(frozen=True)
class Difference:
    workbook: str
    sheet: str
    cell: str
    expected: str
    actual: str
    issue: str


def normalize_value(value: object, *, ignore_case: bool, trim_text: bool) -> object:
    if isinstance(value, str):
        normalized = value.strip() if trim_text else value
        return normalized.casefold() if ignore_case else normalized
    return value


def display_value(value: object) -> str:
    if value is None:
        return ""
    return str(value)


def iter_excel_files(students_dir: Path, template_path: Path) -> Iterable[Path]:
    for path in sorted(students_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        if path.resolve() == template_path.resolve():
            continue
        yield path


def values_match(expected: object, actual: object, args: argparse.Namespace) -> bool:
    return normalize_value(expected, ignore_case=args.ignore_case, trim_text=args.trim_text) == normalize_value(
        actual,
        ignore_case=args.ignore_case,
        trim_text=args.trim_text,
    )


def cell_coordinate(row: int, column: int) -> str:
    return f"{get_column_letter(column)}{row}"


def read_only_non_empty_values(worksheet) -> dict[tuple[int, int], object]:
    values: dict[tuple[int, int], object] = {}
    for row_index, row in enumerate(worksheet.iter_rows(), start=1):
        for column_index, cell in enumerate(row, start=1):
            if cell.value is not None:
                values[(row_index, column_index)] = cell.value
    return values


def editable_non_empty_values(worksheet) -> dict[tuple[int, int], object]:
    values: dict[tuple[int, int], object] = {}
    for position, cell in worksheet._cells.items():
        if cell.value is not None:
            values[position] = cell.value
    return values


def template_value_cache(template_wb) -> dict[str, dict[tuple[int, int], object]]:
    cached = getattr(template_wb, "_assignment_corrector_value_cache", None)
    if cached is not None:
        return cached

    cached = {}
    for sheet_name in template_wb.sheetnames:
        if sheet_name == REPORT_SHEET_NAME:
            continue
        cached[sheet_name] = read_only_non_empty_values(template_wb[sheet_name])
    setattr(template_wb, "_assignment_corrector_value_cache", cached)
    return cached


def add_report_sheet(workbook, differences: list[Difference]) -> None:
    if REPORT_SHEET_NAME in workbook.sheetnames:
        del workbook[REPORT_SHEET_NAME]

    sheet = workbook.create_sheet(REPORT_SHEET_NAME)
    headers = ["Sheet", "Cell", "Issue", "Expected", "Actual"]
    sheet.append(headers)

    for diff in differences:
        sheet.append([diff.sheet, diff.cell, diff.issue, diff.expected, diff.actual])

    for cell in sheet[1]:
        cell.fill = DIFF_FILL

    widths = {
        "A": 24,
        "B": 12,
        "C": 28,
        "D": 45,
        "E": 45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def grade_workbook(template_wb, student_path: Path, corrected_dir: Path, args: argparse.Namespace) -> list[Difference]:
    output_path = corrected_dir / student_path.name
    shutil.copy2(student_path, output_path)

    student_wb = load_workbook(output_path, keep_links=False)
    differences: list[Difference] = []

    try:
        template_values_by_sheet = template_value_cache(template_wb)
        for sheet_name, template_values in template_values_by_sheet.items():
            if sheet_name not in student_wb.sheetnames:
                differences.append(
                    Difference(
                        workbook=student_path.name,
                        sheet=sheet_name,
                        cell="",
                        expected="Sheet exists",
                        actual="Sheet missing",
                        issue="missing sheet",
                    )
                )
                continue

            student_ws = student_wb[sheet_name]
            student_values = editable_non_empty_values(student_ws)

            for row, column in sorted(template_values.keys() | student_values.keys()):
                expected = template_values.get((row, column))
                actual = student_values.get((row, column))

                if values_match(expected, actual, args):
                    continue

                student_cell = student_ws.cell(row=row, column=column)
                student_cell.fill = DIFF_FILL if expected is not None else MISSING_FILL
                differences.append(
                    Difference(
                        workbook=student_path.name,
                        sheet=sheet_name,
                        cell=cell_coordinate(row, column),
                        expected=display_value(expected),
                        actual=display_value(actual),
                        issue="different value",
                    )
                )

        for sheet_name in student_wb.sheetnames:
            if sheet_name not in template_values_by_sheet and sheet_name != REPORT_SHEET_NAME:
                differences.append(
                    Difference(
                        workbook=student_path.name,
                        sheet=sheet_name,
                        cell="",
                        expected="Sheet absent",
                        actual="Extra sheet exists",
                        issue="extra sheet",
                    )
                )

        add_report_sheet(student_wb, differences)
        student_wb.save(output_path)
    finally:
        student_wb.close()
        gc.collect()

    return differences


def write_summary(corrected_dir: Path, differences: list[Difference]) -> Path:
    summary_path = corrected_dir / "grading_summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["workbook", "sheet", "cell", "issue", "expected", "actual"])
        for diff in differences:
            writer.writerow([diff.workbook, diff.sheet, diff.cell, diff.issue, diff.expected, diff.actual])
    return summary_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare student XLSX submissions against a solution template and highlight different cells.",
    )
    parser.add_argument(
        "--students",
        type=Path,
        default=Path("students"),
        help="Folder containing student .xlsx submissions. Default: ./students",
    )
    parser.add_argument(
        "--template",
        type=Path,
        default=None,
        help=f"Template solution workbook. Default search name: {DEFAULT_TEMPLATE_NAME}",
    )
    parser.add_argument(
        "--corrected",
        type=Path,
        default=None,
        help="Output folder for corrected workbooks. Default: <students>/corrected",
    )
    parser.add_argument(
        "--ignore-case",
        action="store_true",
        help="Ignore letter case when comparing text cells.",
    )
    parser.add_argument(
        "--no-trim-text",
        dest="trim_text",
        action="store_false",
        help="Treat leading and trailing spaces as differences.",
    )
    parser.set_defaults(trim_text=True)
    return parser.parse_args()


def resolve_template_path(students_dir: Path, template_arg: Path | None) -> Path:
    if template_arg is not None:
        return template_arg.expanduser().resolve()

    candidates = [
        Path.cwd() / DEFAULT_TEMPLATE_NAME,
        students_dir.parent / DEFAULT_TEMPLATE_NAME,
        students_dir / DEFAULT_TEMPLATE_NAME,
    ]
    for candidate in candidates:
        if candidate.is_file():
            return candidate.resolve()

    return candidates[0].resolve()


def main() -> int:
    args = parse_args()
    students_dir = args.students.expanduser().resolve()
    template_path = resolve_template_path(students_dir, args.template)
    corrected_dir = (args.corrected or students_dir / "corrected").expanduser().resolve()

    if not students_dir.is_dir():
        print(f"Students folder not found: {students_dir}", file=sys.stderr)
        return 1
    if not template_path.is_file():
        print(f"Template workbook not found: {template_path}", file=sys.stderr)
        return 1

    corrected_dir.mkdir(parents=True, exist_ok=True)
    template_wb = load_workbook(template_path, read_only=True, keep_links=False)
    try:
        all_differences: list[Difference] = []
        graded_count = 0

        for student_path in iter_excel_files(students_dir, template_path):
            graded_count += 1
            differences = grade_workbook(template_wb, student_path, corrected_dir, args)
            all_differences.extend(differences)
            print(f"{student_path.name}: {len(differences)} difference(s)")
            gc.collect()
    finally:
        template_wb.close()

    summary_path = write_summary(corrected_dir, all_differences)
    print(f"Graded {graded_count} workbook(s).")
    print(f"Corrected files: {corrected_dir}")
    print(f"Summary: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
