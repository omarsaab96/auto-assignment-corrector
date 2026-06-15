from __future__ import annotations

import csv
import json
import os
import tempfile
import threading
import uuid
from copy import deepcopy
from dataclasses import replace
from pathlib import Path
from types import SimpleNamespace

from flask import Flask, flash, get_flashed_messages, jsonify, redirect, render_template_string, request, session, url_for
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build
from openpyxl import load_workbook
from werkzeug.middleware.proxy_fix import ProxyFix

from grade_assignments import DEFAULT_TEMPLATE_NAME, Difference, grade_workbook
from grade_google_drive import (
    DEFAULT_CORRECTED_FOLDER_PREFIX,
    FOLDER_MIME_TYPE,
    SCOPES,
    XLSX_MIME_TYPE,
    authenticate,
    corrected_destination_folder,
    download_file,
    drive_list,
    escape_drive_query_value,
    find_or_create_corrected_folder,
    find_template_by_name,
    get_file,
    list_student_workbooks,
    resolve_corrected_folder_name,
    timestamped_corrected_folder_name,
    upload_or_replace_file,
    upload_summary_text,
)


SUMMARY_HEADERS = ["workbook", "sheet", "cell", "issue", "expected", "actual"]
SUMMARY_PREVIEW_LIMIT = 200


app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "assignment-corrector-local")
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
app.config["PREFERRED_URL_SCHEME"] = "https" if os.environ.get("RENDER") else "http"
app.config["SESSION_COOKIE_SECURE"] = bool(os.environ.get("RENDER"))
JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()


PAGE = """
<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Assignment Corrector</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f4f6f8;
      --panel: #ffffff;
      --panel-soft: #f8fafc;
      --ink: #111827;
      --muted: #64748b;
      --line: #dbe3ec;
      --line-strong: #cbd5e1;
      --accent: #2563eb;
      --accent-strong: #1d4ed8;
      --accent-soft: #eff6ff;
      --danger: #dc2626;
      --warn: #b45309;
      --success: #15803d;
      --shadow: 0 18px 45px rgba(15, 23, 42, 0.08);
    }

    * { box-sizing: border-box; }

    body {
      margin: 0;
      min-height: 100vh;
      background:
        linear-gradient(180deg, #eef4ff 0, rgba(238, 244, 255, 0) 280px),
        var(--bg);
      color: var(--ink);
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      line-height: 1.45;
    }

    main {
      width: min(1180px, calc(100% - 32px));
      margin: 0 auto;
      padding: 32px 0 52px;
    }

    header {
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 20px;
      margin-bottom: 24px;
    }

    h1 {
      margin: 0 0 4px;
      font-size: 31px;
      line-height: 1.15;
      letter-spacing: 0;
    }

    h2 {
      margin: 0 0 14px;
      font-size: 17px;
      letter-spacing: 0;
    }

    p { margin: 0; color: var(--muted); }

    .header-actions {
      display: flex;
      align-items: center;
      gap: 10px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }

    .grid {
      display: grid;
      grid-template-columns: minmax(0, 1fr) minmax(320px, 390px);
      gap: 20px;
      align-items: start;
    }

    .panel {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 12px;
      padding: 20px;
      box-shadow: var(--shadow);
    }

    label {
      display: block;
      margin-bottom: 15px;
      color: #303946;
      font-size: 13px;
      font-weight: 700;
    }

    input[type="text"] {
      width: 100%;
      min-height: 40px;
      margin-top: 6px;
      border: 1px solid var(--line-strong);
      border-radius: 9px;
      padding: 9px 11px;
      color: var(--ink);
      font: inherit;
      background: #fff;
    }

    input[type="text"]:focus {
      outline: 3px solid rgba(37, 99, 235, 0.14);
      border-color: var(--accent);
    }

    .selected-value {
      display: flex;
      align-items: center;
      gap: 10px;
      min-height: 40px;
      margin-top: 6px;
      border: 1px solid var(--line);
      border-radius: 9px;
      padding: 10px 11px;
      background: var(--panel-soft);
      color: var(--muted);
      overflow-wrap: anywhere;
      font-size: 13px;
    }

    .selected-value strong {
      display: block;
      color: var(--ink);
      font-size: 14px;
    }

    .selected-text {
      min-width: 0;
      overflow-wrap: anywhere;
    }

    .actions {
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      margin-top: 4px;
    }

    button, .button {
      min-height: 40px;
      border: 1px solid var(--accent);
      border-radius: 9px;
      padding: 9px 15px;
      color: #fff;
      background: var(--accent);
      cursor: pointer;
      font: inherit;
      font-weight: 700;
      text-decoration: none;
      transition: background 140ms ease, border-color 140ms ease, transform 140ms ease;
    }

    button.secondary, .button.secondary {
      color: var(--accent);
      background: #fff;
      border-color: var(--line-strong);
    }

    button.small {
      min-height: 32px;
      padding: 6px 10px;
      font-size: 13px;
    }

    button:hover, .button:hover { background: var(--accent-strong); }
    button.secondary:hover, .button.secondary:hover { color: #fff; }
    button:active { transform: translateY(1px); }
    button:disabled {
      cursor: not-allowed;
      opacity: 0.55;
      transform: none;
    }

    .message {
      margin-bottom: 18px;
      border-left: 4px solid var(--danger);
      background: #fff7f7;
      padding: 12px 14px;
      color: #7b2020;
      border-radius: 10px;
    }

    .success {
      border-left-color: var(--accent);
      background: #f0faf6;
      color: #184d3d;
    }

    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }

    th, td {
      padding: 11px 10px;
      border-bottom: 1px solid var(--line);
      text-align: left;
      vertical-align: top;
    }

    th {
      color: #415066;
      font-size: 12px;
      text-transform: uppercase;
      letter-spacing: 0;
      background: var(--panel-soft);
    }

    tr:hover td { background: #fbfdff; }

    .name {
      max-width: 420px;
      overflow-wrap: anywhere;
      font-weight: 700;
    }

    .id {
      max-width: 260px;
      overflow-wrap: anywhere;
      color: var(--muted);
      font-family: Consolas, "Courier New", monospace;
      font-size: 12px;
    }

    .badge {
      display: inline-flex;
      align-items: center;
      min-height: 24px;
      border-radius: 999px;
      padding: 3px 9px;
      background: var(--accent-soft);
      color: var(--accent-strong);
      font-weight: 700;
      font-size: 12px;
    }

    .bad { background: #fff0f0; color: var(--danger); }
    .warn { background: #fff7e6; color: var(--warn); }

    .summary {
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 10px;
      margin-bottom: 16px;
    }

    .metric {
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 12px;
      background: var(--panel-soft);
    }

    .metric strong {
      display: block;
      font-size: 24px;
      line-height: 1.1;
    }

    .metric span { color: var(--muted); font-size: 13px; }

    .browser {
      display: grid;
      gap: 12px;
      max-height: 520px;
      overflow: auto;
    }

    .browser-bar {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 10px;
      border-bottom: 1px solid var(--line);
      padding-bottom: 10px;
    }

    .browser-path {
      min-width: 0;
      color: var(--muted);
      overflow-wrap: anywhere;
      font-size: 13px;
    }

    .row-actions {
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
    }

    .drive-name {
      display: flex;
      align-items: center;
      gap: 10px;
      min-width: 0;
    }

    .drive-icon {
      width: 22px;
      height: 22px;
      flex: 0 0 22px;
    }

    .folder-icon { color: #f59e0b; }
    .file-icon { color: var(--success); }
    .rename-icon { color: var(--accent); }

    .drive-label {
      min-width: 0;
      overflow-wrap: anywhere;
    }

    .folder-link {
      appearance: none;
      border: 0;
      background: transparent;
      color: var(--ink);
      cursor: pointer;
      font: inherit;
      font-weight: 700;
      padding: 0;
      text-align: left;
      overflow-wrap: anywhere;
    }

    .folder-link:hover {
      color: var(--accent);
      text-decoration: underline;
    }

    .muted { color: var(--muted); }

    .hidden { display: none !important; }

    .progress-panel {
      margin-top: 20px;
    }

    .progress-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      margin-bottom: 16px;
    }

    .loader {
      width: 28px;
      height: 28px;
      border: 3px solid var(--line);
      border-top-color: var(--accent);
      border-radius: 50%;
      animation: spin 900ms linear infinite;
      flex: 0 0 28px;
    }

    @keyframes spin {
      to { transform: rotate(360deg); }
    }

    .progress-track {
      height: 12px;
      overflow: hidden;
      border-radius: 999px;
      background: #e8eef6;
      border: 1px solid var(--line);
    }

    .progress-fill {
      width: 0%;
      height: 100%;
      background: var(--accent);
      transition: width 180ms ease;
    }

    .progress-meta {
      display: flex;
      justify-content: space-between;
      gap: 10px;
      margin: 8px 0 16px;
      color: var(--muted);
      font-size: 13px;
    }

    .submission-list {
      display: grid;
      gap: 8px;
      margin-top: 14px;
    }

    .submission-row {
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 10px;
      align-items: center;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px 12px;
      background: var(--panel-soft);
    }

    .submission-name {
      min-width: 0;
      overflow-wrap: anywhere;
      font-weight: 700;
    }

    .status-pill {
      border-radius: 999px;
      padding: 4px 9px;
      background: var(--accent-soft);
      color: var(--accent-strong);
      font-size: 12px;
      font-weight: 800;
    }

    .status-pending { background: #f1f5f9; color: var(--muted); }
    .status-running { background: var(--accent-soft); color: var(--accent-strong); }
    .status-done { background: #ecfdf5; color: var(--success); }
    .status-error { background: #fef2f2; color: var(--danger); }

    .setup-summary {
      display: grid;
      gap: 12px;
    }

    .summary-row {
      display: flex;
      align-items: flex-start;
      gap: 11px;
      border-bottom: 1px solid var(--line);
      padding-bottom:10px;
      # border-radius: 10px;
      # padding: 12px;
      # background: var(--panel-soft);
    }

    .summary-row-body {
      min-width: 0;
    }

    .summary-row-body strong{
      font-size:12px;
    }

    .summary-row-title {
      display: block;
      margin-bottom: 2px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 800;
      text-transform: uppercase;
    }

    .summary-row strong {
      display: block;
      color: var(--ink);
      overflow-wrap: anywhere;
    }

    .summary-row span:last-child {
      display: block;
      color: var(--muted);
      font-size: 12px;
      overflow-wrap: anywhere;
    }

    .run-options {
      display: grid;
      gap: 11px;
      margin-top: 4px;
    }

    .output-name {
      margin-top: 4px;
    }

    .toggle-row {
      display: flex;
      align-items: center;
      gap: 10px;
      color: var(--ink);
      font-weight: 700;
      cursor: pointer;
    }

    .toggle-row input {
      width: 18px;
      height: 18px;
      margin: 0;
      accent-color: var(--accent);
    }

    @media (max-width: 880px) {
      header, .grid { display: block; }
      .panel { margin-bottom: 16px; }
      .summary { grid-template-columns: 1fr; }
    }
  </style>
</head>
<body>
  <main>
    <header>
      <div>
        <h1>Assignment Corrector</h1>
        <p>Compare each student workbook with the solution file and upload marked copies to Drive.</p>
      </div>
      <div class="header-actions">
        {% if signed_in %}
          <span class="badge">Google Drive connected</span>
          <a class="button secondary" href="{{ url_for('logout') }}">Sign out</a>
        {% else %}
          <a class="button" href="{{ url_for('login') }}">Sign in with Google</a>
        {% endif %}
      </div>
    </header>

    {% if error %}
      <div class="message">{{ error }}</div>
    {% endif %}
    {% if success_message %}
      <div class="message success">{{ success_message }}</div>
    {% endif %}
    {% if result %}
      <div class="message success">
        Corrected {{ result.graded_count }} workbook(s). Output folder ID:
        <span class="id">{{ result.corrected_folder_id }}</span>
      </div>
    {% endif %}

    <div id="work-area" class="grid">
      <section class="panel">
        <div class="browser">
          <div class="browser-bar">
            <div>
              <strong id="browser-title">My Drive</strong>
              <div id="browser-path" class="browser-path">root</div>
            </div>
            <div class="actions">
              <button type="button" id="browser-up" class="secondary small">Back</button>
              <button type="button" id="browser-refresh" class="secondary small">Refresh</button>
            </div>
          </div>
          <table>
            <thead>
              <tr>
                <th>Name</th>
                <th>Type</th>
                <th>Action</th>
              </tr>
            </thead>
            <tbody id="browser-body">
              <tr><td colspan="3" class="muted">Loading Drive files...</td></tr>
            </tbody>
          </table>
        </div>
      </section>

      <section class="panel">
        <h2>Correction Summary</h2>
        <form id="correction-form" method="post" autocomplete="off">
          <input id="students-folder-id" type="hidden" name="students_folder_id" value="{{ form.students_folder_id }}" required autocomplete="off">
          <input id="students-folder-name-input" type="hidden" name="students_folder_name" value="{{ form.students_folder_name }}" autocomplete="off">
          <input id="template-file-id" type="hidden" name="template_file_id" value="{{ form.template_file_id }}" autocomplete="off">
          <input id="template-file-name-input" type="hidden" name="template_file_name" value="{{ form.template_file_name }}" autocomplete="off">
          <input type="hidden" name="template_name" value="{{ form.template_name }}" autocomplete="off">

          <div class="setup-summary">
            <div class="summary-row">
              <span id="students-folder-icon"></span>
              <div class="summary-row-body">
                <span class="summary-row-title">Students Folder</span>
                <strong id="students-folder-name">{{ form.students_folder_name or "No folder selected" }}</strong>
              </div>
            </div>

            <div class="summary-row">
              <span id="template-file-icon"></span>
              <div class="summary-row-body">
                <span class="summary-row-title">Solution Workbook</span>
                <strong id="template-file-name">{{ form.template_file_name or "No file selected" }}</strong>
              </div>
            </div>

            <div class="summary-row">
              <span id="output-folder-name-icon"></span>
              <div class="summary-row-body">
                <span class="summary-row-title">Output folder</span>
                <input id="corrected-folder-name" type="text" name="corrected_folder_name" value="{{ form.corrected_folder_name }}" placeholder="{{ corrected_folder_placeholder }}" autocomplete="off">
              </div>
            </div>

            <div class="summary-row">
              <span id="latest-only-icon"></span>
              <div class="summary-row-body">
                <span class="summary-row-title">Submission mode</span>
                <label class="toggle-row">
                  <input id="latest-only" type="checkbox" name="latest_only" value="1" {% if form.latest_only %}checked{% endif %}>
                  Correct latest submission only
                </label>
              </div>
            </div>

            <div class="run-options">
              <div class="actions">
                <button type="submit" name="action" value="grade">Start correcting</button>
              </div>
            </div>
          </div>
        </form>
      </section>
    </div>

    <section id="progress-panel" class="panel progress-panel hidden">
      <div class="progress-head">
        <div>
          <h2 id="progress-title">Correcting submissions</h2>
          <p id="progress-stage">Preparing correction run...</p>
        </div>
        <div class="actions">
          <button type="button" id="cancel-correction" class="secondary small">Cancel</button>
          <div id="progress-loader" class="loader"></div>
        </div>
      </div>

      <div class="progress-track">
        <div id="overall-progress-fill" class="progress-fill"></div>
      </div>
      <div class="progress-meta">
        <span id="overall-progress-label">0%</span>
        <span id="overall-progress-count">0 of 0 complete</span>
      </div>

      <div id="progress-message" class="message hidden"></div>
      <div id="submission-list" class="submission-list"></div>
    </section>

    {% if result and result.differences %}
      <section class="panel" style="margin-top:18px;">
        <div class="summary">
          <div class="metric"><strong>{{ result.graded_count }}</strong><span>workbooks corrected</span></div>
          <div class="metric"><strong>{{ result.total_differences }}</strong><span>different cells or sheets</span></div>
          <div class="metric"><strong>{{ result.corrected_folder_name }}</strong><span>Drive output folder</span></div>
        </div>

        <h2>Correction Report</h2>
        <table>
          <thead>
            <tr>
              <th>Workbook</th>
              <th>Sheet</th>
              <th>Cell</th>
              <th>Issue</th>
              <th>Expected</th>
              <th>Actual</th>
            </tr>
          </thead>
          <tbody>
            {% for diff in result.differences[:200] %}
              <tr>
                <td class="name">{{ diff.workbook }}</td>
                <td>{{ diff.sheet }}</td>
                <td>{{ diff.cell }}</td>
                <td><span class="badge warn">{{ diff.issue }}</span></td>
                <td>{{ diff.expected }}</td>
                <td>{{ diff.actual }}</td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </section>
    {% endif %}
  </main>
  <script>
    const folderMimeType = "{{ folder_mime_type }}";
    const xlsxMimeType = "{{ xlsx_mime_type }}";
    const resetSummaryOnLoad = {{ "true" if reset_summary_on_load else "false" }};
    const folderStack = [{ id: "root", name: "My Drive" }];

    const browserBody = document.getElementById("browser-body");
    const browserTitle = document.getElementById("browser-title");
    const browserPath = document.getElementById("browser-path");
    const browserUp = document.getElementById("browser-up");
    const browserRefresh = document.getElementById("browser-refresh");
    const workArea = document.getElementById("work-area");
    const correctionForm = document.getElementById("correction-form");
    const progressPanel = document.getElementById("progress-panel");
    const progressTitle = document.getElementById("progress-title");
    const progressStage = document.getElementById("progress-stage");
    const progressLoader = document.getElementById("progress-loader");
    const progressFill = document.getElementById("overall-progress-fill");
    const progressLabel = document.getElementById("overall-progress-label");
    const progressCount = document.getElementById("overall-progress-count");
    const progressMessage = document.getElementById("progress-message");
    const submissionList = document.getElementById("submission-list");
    const cancelCorrection = document.getElementById("cancel-correction");
    const studentsFolderId = document.getElementById("students-folder-id");
    const studentsFolderName = document.getElementById("students-folder-name");
    const studentsFolderDetail = document.getElementById("students-folder-detail");
    const studentsFolderNameInput = document.getElementById("students-folder-name-input");
    const studentsFolderIcon = document.getElementById("students-folder-icon");
    const templateFileId = document.getElementById("template-file-id");
    const templateFileName = document.getElementById("template-file-name");
    const templateFileDetail = document.getElementById("template-file-detail");
    const templateFileNameInput = document.getElementById("template-file-name-input");
    const templateFileIcon = document.getElementById("template-file-icon");
    const outputFolderNameIcon = document.getElementById("output-folder-name-icon");
    const correctedFolderName = document.getElementById("corrected-folder-name");
    const latestOnlyIcon = document.getElementById("latest-only-icon");
    let activeJobId = null;
    let correctionCancelled = false;

    function currentFolder() {
      return folderStack[folderStack.length - 1];
    }

    function updatePath() {
      browserTitle.textContent = currentFolder().name;
      browserPath.textContent = folderStack.map((folder) => folder.name).join(" / ");
      browserUp.disabled = folderStack.length <= 1;
    }

    function renderLoading() {
      browserBody.innerHTML = '<tr><td colspan="3" class="muted">Loading Drive files...</td></tr>';
    }

    function renderError(message) {
      browserBody.innerHTML = `<tr><td colspan="3" class="muted">${message}</td></tr>`;
    }

    function button(label, className, onClick) {
      const element = document.createElement("button");
      element.type = "button";
      element.className = className;
      element.textContent = label;
      element.addEventListener("click", onClick);
      return element;
    }

    function iconSvg(type) {
      if (type === "folder") {
        return `
          <svg class="drive-icon folder-icon" viewBox="0 0 24 24" fill="none" aria-hidden="true">
            <path d="M3 6.75A2.75 2.75 0 0 1 5.75 4h4.1c.73 0 1.43.29 1.94.81l1.4 1.4c.24.24.57.39.92.39h4.14A2.75 2.75 0 0 1 21 9.35v7.9A2.75 2.75 0 0 1 18.25 20H5.75A2.75 2.75 0 0 1 3 17.25V6.75Z" fill="currentColor" opacity=".18"/>
            <path d="M3 8.25A2.75 2.75 0 0 1 5.75 5.5h3.74c.6 0 1.17.24 1.59.66l1.26 1.26c.42.42.99.66 1.59.66h4.32A2.75 2.75 0 0 1 21 10.83v6.42A2.75 2.75 0 0 1 18.25 20H5.75A2.75 2.75 0 0 1 3 17.25v-9Z" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/>
          </svg>`;
      }
      if (type === "rename") {
        return `
          <svg class="drive-icon rename-icon" viewBox="0 0 24 24" fill="none" aria-hidden="true">
            <path d="M4 17.75V20h2.25L17.8 8.45l-2.25-2.25L4 17.75Z" fill="currentColor" opacity=".16"/>
            <path d="M4 17.75V20h2.25L18.2 8.05a1.6 1.6 0 0 0 0-2.26l-.99-.99a1.6 1.6 0 0 0-2.26 0L4 17.75Z" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/>
            <path d="M14.25 5.5l4.25 4.25" stroke="currentColor" stroke-width="1.7" stroke-linecap="round"/>
            <path d="M4 21h16" stroke="currentColor" stroke-width="1.7" stroke-linecap="round"/>
          </svg>`;
      }
      if (type === "latest") {
        return `
          <svg class="drive-icon rename-icon" viewBox="0 0 24 24" fill="none" aria-hidden="true">
            <path d="M12 5.25a6.75 6.75 0 1 0 6.75 6.75" stroke="currentColor" stroke-width="1.7" stroke-linecap="round"/>
            <path d="M12 8v4.25l2.75 1.65" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"/>
            <path d="M17.5 3.75h3v3" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"/>
            <path d="M20.5 3.75 16.75 7.5" stroke="currentColor" stroke-width="1.7" stroke-linecap="round"/>
          </svg>`;
      }
      return `
        <svg class="drive-icon file-icon" viewBox="0 0 24 24" fill="none" aria-hidden="true">
          <path d="M6.75 3.5h7.15L19.5 9.1v8.15A3.25 3.25 0 0 1 16.25 20.5h-9.5A2.25 2.25 0 0 1 4.5 18.25V5.75A2.25 2.25 0 0 1 6.75 3.5Z" fill="currentColor" opacity=".13"/>
          <path d="M14 3.75v3.5A1.75 1.75 0 0 0 15.75 9H19" stroke="currentColor" stroke-width="1.7" stroke-linecap="round" stroke-linejoin="round"/>
          <path d="M6.75 3.5h7.15L19.5 9.1v8.15A3.25 3.25 0 0 1 16.25 20.5h-9.5A2.25 2.25 0 0 1 4.5 18.25V5.75A2.25 2.25 0 0 1 6.75 3.5Z" stroke="currentColor" stroke-width="1.7" stroke-linejoin="round"/>
          <path d="M8 13h8M8 16h5" stroke="currentColor" stroke-width="1.7" stroke-linecap="round"/>
        </svg>`;
    }

    function clearCorrectionSummary() {
      studentsFolderId.value = "";
      studentsFolderNameInput.value = "";
      studentsFolderName.textContent = "No folder selected";
      if (studentsFolderDetail) {
        studentsFolderDetail.textContent = "Select a folder from Drive";
      }
      templateFileId.value = "";
      templateFileNameInput.value = "";
      templateFileName.textContent = "No file selected";
      if (templateFileDetail) {
        templateFileDetail.textContent = "Select an .xlsx file from Drive";
      }
      correctedFolderName.value = "";
    }

    function restoreCorrectionSetup() {
      correctionCancelled = true;
      progressPanel.classList.add("hidden");
      workArea.classList.remove("hidden");
      progressLoader.classList.add("hidden");
      cancelCorrection.disabled = false;
    }

    function renderItems(items) {
      browserBody.innerHTML = "";
      if (!items.length) {
        renderError("No folders or Excel files found here.");
        return;
      }

      for (const item of items) {
        const row = document.createElement("tr");
        row.className = item.mimeType === folderMimeType ? "folder-row" : "file-row";

        const nameCell = document.createElement("td");
        nameCell.className = "name";
        const nameWrap = document.createElement("div");
        nameWrap.className = "drive-name";
        nameWrap.innerHTML = iconSvg(item.mimeType === folderMimeType ? "folder" : "file");
        const label = document.createElement(item.mimeType === folderMimeType ? "button" : "span");
        label.className = item.mimeType === folderMimeType ? "drive-label folder-link" : "drive-label";
        label.textContent = item.name;
        if (item.mimeType === folderMimeType) {
          label.type = "button";
          label.addEventListener("click", () => {
            folderStack.push({ id: item.id, name: item.name });
            loadFolder(item.id);
          });
        }
        nameWrap.appendChild(label);
        nameCell.appendChild(nameWrap);

        const typeCell = document.createElement("td");
        typeCell.textContent = item.mimeType === folderMimeType ? "Folder" : "Excel workbook";

        const actionsCell = document.createElement("td");
        const actions = document.createElement("div");
        actions.className = "row-actions";

        if (item.mimeType === folderMimeType) {
          actions.appendChild(button("Use for students", "small", () => {
            studentsFolderId.value = item.id;
            studentsFolderNameInput.value = item.name;
            studentsFolderName.textContent = item.name;
            if (studentsFolderDetail) {
              studentsFolderDetail.textContent = item.id;
            }
            studentsFolderIcon.innerHTML = iconSvg("folder");
          }));
        } else {
          actions.appendChild(button("Use as solution", "small", () => {
            templateFileId.value = item.id;
            templateFileNameInput.value = item.name;
            templateFileName.textContent = item.name;
            if (templateFileDetail) {
              templateFileDetail.textContent = item.id;
            }
            templateFileIcon.innerHTML = iconSvg("file");
          }));
        }

        actionsCell.appendChild(actions);
        row.appendChild(nameCell);
        row.appendChild(typeCell);
        row.appendChild(actionsCell);
        browserBody.appendChild(row);
      }
    }

    async function loadFolder(folderId) {
      updatePath();
      renderLoading();
      try {
        const response = await fetch(`/api/drive/children?folder_id=${encodeURIComponent(folderId)}`);
        const payload = await response.json();
        if (!response.ok) {
          throw new Error(payload.error || "Unable to load Drive folder.");
        }
        renderItems(payload.items);
      } catch (error) {
        renderError(error.message);
      }
    }

    browserUp.addEventListener("click", () => {
      if (folderStack.length > 1) {
        folderStack.pop();
        loadFolder(currentFolder().id);
      }
    });

    browserRefresh.addEventListener("click", () => loadFolder(currentFolder().id));

    function submissionStatusClass(status) {
      if (status === "done") return "status-done";
      if (status === "running") return "status-running";
      if (status === "error") return "status-error";
      return "status-pending";
    }

    function showProgressMessage(kind, text) {
      progressMessage.className = `message ${kind === "success" ? "success" : ""}`;
      progressMessage.textContent = text;
      progressMessage.classList.remove("hidden");
    }

    function renderJob(job) {
      const total = job.total || job.submissions.length || 0;
      const completed = job.completed || 0;
      const percent = total ? Math.round((completed / total) * 100) : (job.status === "running" ? 8 : 0);

      progressTitle.textContent = job.status === "error" ? "Correction failed" : job.status === "done" ? "Correction complete" : "Correcting submissions";
      progressStage.textContent = job.stage || "Working...";
      progressFill.style.width = `${percent}%`;
      progressLabel.textContent = `${percent}%`;
      progressCount.textContent = `${completed} of ${total} complete`;

      submissionList.innerHTML = "";
      for (const submission of job.submissions) {
        const row = document.createElement("div");
        row.className = "submission-row";

        const name = document.createElement("div");
        name.className = "submission-name";
        name.textContent = submission.name;

        const status = document.createElement("span");
        status.className = `status-pill ${submissionStatusClass(submission.status)}`;
        status.textContent = submission.status === "done"
          ? `${submission.differences || 0} difference(s)`
          : submission.status;

        row.appendChild(name);
        row.appendChild(status);
        submissionList.appendChild(row);
      }

      if (job.status === "cancelled") {
        progressLoader.classList.add("hidden");
        cancelCorrection.disabled = true;
        showProgressMessage("error", "Correction cancelled.");
      } else if (job.status === "done") {
        progressFill.style.width = "100%";
        progressLabel.textContent = "100%";
        progressLoader.classList.add("hidden");
        cancelCorrection.disabled = true;
        showProgressMessage("success", `Correction finished. Output folder: ${job.corrected_folder_name}`);
      } else if (job.status === "error") {
        progressLoader.classList.add("hidden");
        cancelCorrection.disabled = true;
        showProgressMessage("error", job.error || "Correction failed.");
      }
    }

    async function pollJob(jobId) {
      const response = await fetch(`/api/correction/status/${encodeURIComponent(jobId)}`);
      const job = await response.json();
      if (!response.ok) {
        throw new Error(job.error || "Unable to read correction progress.");
      }

      renderJob(job);
      if (!correctionCancelled && (job.status === "running" || job.status === "queued")) {
        window.setTimeout(() => pollJob(jobId).catch((error) => {
          progressLoader.classList.add("hidden");
          showProgressMessage("error", error.message);
        }), 900);
      }
    }

    correctionForm.addEventListener("submit", async (event) => {
      event.preventDefault();
      correctionCancelled = false;
      activeJobId = null;
      progressMessage.classList.add("hidden");
      progressLoader.classList.remove("hidden");
      cancelCorrection.disabled = false;
      workArea.classList.add("hidden");
      progressPanel.classList.remove("hidden");
      submissionList.innerHTML = "";
      progressFill.style.width = "0%";
      progressLabel.textContent = "0%";
      progressCount.textContent = "0 of 0 complete";
      progressStage.textContent = "Starting correction run...";

      try {
        const response = await fetch("/api/correction/start", {
          method: "POST",
          body: new FormData(correctionForm),
        });
        const payload = await response.json();
        if (!response.ok) {
          throw new Error(payload.error || "Unable to start correction.");
        }
        activeJobId = payload.job_id;
        await pollJob(payload.job_id);
      } catch (error) {
        progressLoader.classList.add("hidden");
        showProgressMessage("error", error.message);
      }
    });

    cancelCorrection.addEventListener("click", async () => {
      const jobId = activeJobId;
      restoreCorrectionSetup();
      if (jobId) {
        try {
          await fetch(`/api/correction/cancel/${encodeURIComponent(jobId)}`, { method: "POST" });
        } catch (_) {
        }
      }
    });

    function isReloadNavigation() {
      const entries = performance.getEntriesByType ? performance.getEntriesByType("navigation") : [];
      if (entries.length) {
        return entries[0].type === "reload";
      }
      return performance.navigation && performance.navigation.type === performance.navigation.TYPE_RELOAD;
    }

    function scheduleCorrectionSummaryReset() {
      clearCorrectionSummary();
      window.setTimeout(clearCorrectionSummary, 0);
      window.setTimeout(clearCorrectionSummary, 150);
    }

    loadFolder("root");
    studentsFolderIcon.innerHTML = iconSvg("folder");
    templateFileIcon.innerHTML = iconSvg("file");
    outputFolderNameIcon.innerHTML = iconSvg("rename");
    latestOnlyIcon.innerHTML = iconSvg("latest");
    if (resetSummaryOnLoad || isReloadNavigation()) {
      scheduleCorrectionSummaryReset();
    }

    window.addEventListener("pageshow", (event) => {
      if (event.persisted || resetSummaryOnLoad || isReloadNavigation()) {
        scheduleCorrectionSummaryReset();
      }
    });

    window.addEventListener("beforeunload", () => {
      clearCorrectionSummary();
    });
  </script>
</body>
</html>
"""


def default_form() -> dict[str, object]:
    return {
        "students_folder_id": "",
        "students_folder_name": "",
        "template_file_id": "",
        "template_file_name": "",
        "template_name": DEFAULT_TEMPLATE_NAME,
        "corrected_folder_name": "",
        "latest_only": False,
        "ignore_case": False,
        "trim_text": True,
    }


def form_from_request() -> dict[str, object]:
    return {
        "students_folder_id": request.form.get("students_folder_id", "").strip(),
        "students_folder_name": request.form.get("students_folder_name", "").strip(),
        "template_file_id": request.form.get("template_file_id", "").strip(),
        "template_file_name": request.form.get("template_file_name", "").strip(),
        "template_name": request.form.get("template_name", DEFAULT_TEMPLATE_NAME).strip() or DEFAULT_TEMPLATE_NAME,
        "corrected_folder_name": request.form.get("corrected_folder_name", "").strip(),
        "latest_only": request.form.get("latest_only") == "1",
        "ignore_case": False,
        "trim_text": True,
    }


def credentials_to_dict(credentials: Credentials) -> dict:
    return {
        "token": credentials.token,
        "refresh_token": credentials.refresh_token,
        "token_uri": credentials.token_uri,
        "client_id": credentials.client_id,
        "client_secret": credentials.client_secret,
        "scopes": credentials.scopes,
    }


def load_google_client_config() -> dict:
    raw_config = os.environ.get("GOOGLE_OAUTH_CLIENT_JSON")
    if raw_config:
        return json.loads(raw_config)

    credentials_path = Path("credentials.json")
    if credentials_path.is_file():
        return json.loads(credentials_path.read_text(encoding="utf-8"))

    raise FileNotFoundError("Google OAuth config not found. Set GOOGLE_OAUTH_CLIENT_JSON on Render.")


def make_oauth_flow(state: str | None = None) -> Flow:
    flow = Flow.from_client_config(load_google_client_config(), scopes=SCOPES, state=state)
    flow.redirect_uri = url_for("oauth2callback", _external=True)
    return flow


def get_drive_service(credentials_data: dict | None = None):
    if credentials_data is None and "credentials" in session:
        credentials_data = session["credentials"]

    if credentials_data:
        credentials = Credentials(**credentials_data)
        if credentials.expired and credentials.refresh_token:
            credentials.refresh(Request())
            if "credentials" in session:
                session["credentials"] = credentials_to_dict(credentials)
        if not credentials.valid:
            raise RuntimeError("Google Drive login expired. Sign in again.")
        return build("drive", "v3", credentials=credentials)

    if os.environ.get("RENDER"):
        raise RuntimeError("Sign in with Google Drive first.")

    return authenticate(Path("credentials.json").resolve(), Path("token.json").resolve())


def current_credentials_data() -> dict | None:
    if "credentials" in session:
        return dict(session["credentials"])
    if os.environ.get("RENDER"):
        return None
    return None


def filter_latest_submissions(students: list[dict]) -> list[dict]:
    latest_by_folder: dict[tuple[str, ...], dict] = {}
    direct_files: list[dict] = []

    for student in students:
        folder_path = tuple(student.get("folder_path") or [])
        if not folder_path:
            direct_files.append(student)
            continue

        current = latest_by_folder.get(folder_path)
        if current is None or student.get("modifiedTime", "") > current.get("modifiedTime", ""):
            latest_by_folder[folder_path] = student

    latest_files = sorted(latest_by_folder.values(), key=lambda item: item["relative_path"].casefold())
    return sorted(direct_files, key=lambda item: item["relative_path"].casefold()) + latest_files


def load_drive_files(service, form: dict[str, object]):
    excluded_folder_names = {DEFAULT_CORRECTED_FOLDER_PREFIX}
    if form["corrected_folder_name"]:
        excluded_folder_names.add(str(form["corrected_folder_name"]))

    template = (
        get_file(service, str(form["template_file_id"]))
        if form["template_file_id"]
        else find_template_by_name(service, str(form["template_name"]))
    )
    students = list_student_workbooks(
        service,
        str(form["students_folder_id"]),
        template["id"],
        excluded_folder_names=excluded_folder_names,
    )
    if form["latest_only"]:
        students = filter_latest_submissions(students)
    return template, students


def create_summary_file(corrected_dir: Path) -> Path:
    summary_path = corrected_dir / "grading_summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(SUMMARY_HEADERS)
    return summary_path


def append_summary_rows(summary_path: Path, workbook_name: str, differences: list[Difference]) -> None:
    with summary_path.open("a", newline="", encoding="utf-8") as csv_file:
        writer = csv.writer(csv_file)
        for diff in differences:
            writer.writerow([workbook_name, diff.sheet, diff.cell, diff.issue, diff.expected, diff.actual])


def add_preview_differences(
    preview: list[Difference],
    workbook_name: str,
    differences: list[Difference],
) -> None:
    remaining = SUMMARY_PREVIEW_LIMIT - len(preview)
    if remaining <= 0:
        return
    preview.extend(replace(diff, workbook=workbook_name) for diff in differences[:remaining])


def grade_drive_files(service, form: dict[str, object], template: dict, students: list[dict]) -> dict[str, object]:
    corrected_folder_name = resolve_corrected_folder_name(str(form["corrected_folder_name"]))
    corrected_folder_id = find_or_create_corrected_folder(
        service,
        str(form["students_folder_id"]),
        corrected_folder_name,
    )

    with tempfile.TemporaryDirectory(prefix="assignment-corrector-ui-") as temp_dir_name:
        temp_dir = Path(temp_dir_name)
        students_dir = temp_dir / "students"
        corrected_dir = temp_dir / "corrected"
        students_dir.mkdir()
        corrected_dir.mkdir()

        template_path = temp_dir / template["name"]
        download_file(service, template["id"], template_path)
        template_wb = load_workbook(template_path)
        grade_args = SimpleNamespace(ignore_case=form["ignore_case"], trim_text=form["trim_text"])

        preview_differences: list[Difference] = []
        difference_counts: dict[str, int] = {}
        total_differences = 0
        summary_path = create_summary_file(corrected_dir)

        for student in students:
            student_source_dir = students_dir / student["id"]
            student_source_dir.mkdir()
            student_path = student_source_dir / student["name"]
            download_file(service, student["id"], student_path)
            student_corrected_dir = corrected_dir / student["id"]
            student_corrected_dir.mkdir()
            differences = grade_workbook(template_wb, student_path, student_corrected_dir, grade_args)
            difference_counts[student["relative_path"]] = len(differences)
            total_differences += len(differences)
            append_summary_rows(summary_path, student["relative_path"], differences)
            add_preview_differences(preview_differences, student["relative_path"], differences)

            corrected_path = student_corrected_dir / student_path.name
            destination_folder_id = corrected_destination_folder(service, corrected_folder_id, student)
            upload_or_replace_file(service, destination_folder_id, corrected_path, XLSX_MIME_TYPE)

        upload_summary_text(service, corrected_folder_id, summary_path)

    return {
        "graded_count": len(students),
        "corrected_folder_id": corrected_folder_id,
        "corrected_folder_name": corrected_folder_name,
        "differences": preview_differences,
        "difference_counts": difference_counts,
        "total_differences": total_differences,
    }


def job_snapshot(job_id: str) -> dict | None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        return deepcopy(job) if job else None


def update_job(job_id: str, **updates) -> None:
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(updates)


def update_submission(job_id: str, index: int, **updates) -> None:
    with JOBS_LOCK:
        if job_id in JOBS and 0 <= index < len(JOBS[job_id]["submissions"]):
            JOBS[job_id]["submissions"][index].update(updates)


def job_cancel_requested(job_id: str) -> bool:
    with JOBS_LOCK:
        return bool(JOBS.get(job_id, {}).get("cancel_requested"))


def trash_drive_file(service, file_id: str) -> None:
    service.files().update(fileId=file_id, body={"trashed": True}, fields="id").execute()


def run_correction_job(job_id: str, form: dict[str, object], credentials_data: dict | None) -> None:
    current_index = None
    corrected_folder_id = ""
    try:
        update_job(job_id, status="running", stage="Connecting to Google Drive...")
        service = get_drive_service(credentials_data)

        update_job(job_id, stage="Finding solution workbook and submissions...")
        template, students = load_drive_files(service, form)
        corrected_folder_name = resolve_corrected_folder_name(str(form["corrected_folder_name"]))

        submissions = [
            {
                "name": student["relative_path"],
                "status": "pending",
                "differences": 0,
            }
            for student in students
        ]
        update_job(
            job_id,
            total=len(students),
            submissions=submissions,
            corrected_folder_name=corrected_folder_name,
            stage="Downloading solution workbook...",
        )

        with tempfile.TemporaryDirectory(prefix="assignment-corrector-ui-") as temp_dir_name:
            temp_dir = Path(temp_dir_name)
            students_dir = temp_dir / "students"
            corrected_dir = temp_dir / "corrected"
            students_dir.mkdir()
            corrected_dir.mkdir()

            template_path = temp_dir / template["name"]
            download_file(service, template["id"], template_path)
            template_wb = load_workbook(template_path)
            grade_args = SimpleNamespace(ignore_case=form["ignore_case"], trim_text=form["trim_text"])

            corrected_paths: list[tuple[dict, Path]] = []
            total_differences = 0
            summary_path = create_summary_file(corrected_dir)

            for index, student in enumerate(students):
                if job_cancel_requested(job_id):
                    update_job(job_id, status="cancelled", stage="Correction cancelled.")
                    return

                current_index = index
                update_submission(job_id, index, status="running")
                update_job(job_id, stage=f"Correcting {student['relative_path']}...")

                student_source_dir = students_dir / student["id"]
                student_source_dir.mkdir()
                student_path = student_source_dir / student["name"]
                download_file(service, student["id"], student_path)

                student_corrected_dir = corrected_dir / student["id"]
                student_corrected_dir.mkdir()
                differences = grade_workbook(template_wb, student_path, student_corrected_dir, grade_args)
                total_differences += len(differences)
                append_summary_rows(summary_path, student["relative_path"], differences)
                corrected_path = student_corrected_dir / student_path.name
                corrected_paths.append((student, corrected_path))

                with JOBS_LOCK:
                    JOBS[job_id]["completed"] += 1
                    JOBS[job_id]["total_differences"] = total_differences
                update_submission(job_id, index, status="done", differences=len(differences))

            if job_cancel_requested(job_id):
                update_job(job_id, status="cancelled", stage="Correction cancelled.")
                return

            update_job(job_id, stage="Preparing Drive output folder...")
            corrected_folder_id = find_or_create_corrected_folder(
                service,
                str(form["students_folder_id"]),
                corrected_folder_name,
            )
            update_job(job_id, corrected_folder_id=corrected_folder_id)

            try:
                for student, corrected_path in corrected_paths:
                    if job_cancel_requested(job_id):
                        trash_drive_file(service, corrected_folder_id)
                        update_job(
                            job_id,
                            status="cancelled",
                            corrected_folder_id="",
                            stage="Correction cancelled. No corrected folder was kept.",
                        )
                        return

                    update_job(job_id, stage=f"Uploading {student['relative_path']}...")
                    destination_folder_id = corrected_destination_folder(service, corrected_folder_id, student)
                    upload_or_replace_file(service, destination_folder_id, corrected_path, XLSX_MIME_TYPE)

                if job_cancel_requested(job_id):
                    trash_drive_file(service, corrected_folder_id)
                    update_job(
                        job_id,
                        status="cancelled",
                        corrected_folder_id="",
                        stage="Correction cancelled. No corrected folder was kept.",
                    )
                    return

                update_job(job_id, stage="Uploading summary report...")
                upload_summary_text(service, corrected_folder_id, summary_path)
            except Exception:
                if job_cancel_requested(job_id) and corrected_folder_id:
                    trash_drive_file(service, corrected_folder_id)
                    update_job(
                        job_id,
                        status="cancelled",
                        corrected_folder_id="",
                        stage="Correction cancelled. No corrected folder was kept.",
                    )
                    return
                raise

        update_job(
            job_id,
            status="done",
            stage="Correction complete.",
            error=None,
            total_differences=total_differences,
        )
    except Exception as exc:
        if current_index is not None:
            update_submission(job_id, current_index, status="error")
        update_job(job_id, status="error", stage="Correction failed.", error=str(exc))


def list_drive_browser_items(service, folder_id: str) -> list[dict[str, str]]:
    safe_folder_id = escape_drive_query_value(folder_id)
    query = (
        f"'{safe_folder_id}' in parents and "
        f"(mimeType = '{FOLDER_MIME_TYPE}' or mimeType = '{XLSX_MIME_TYPE}') and "
        "trashed = false"
    )
    items = drive_list(service, query, "id, name, mimeType")
    return sorted(
        items,
        key=lambda item: (
            item["mimeType"] != FOLDER_MIME_TYPE,
            item["name"].casefold(),
        ),
    )


@app.route("/", methods=["GET", "POST"])
def index():
    form = default_form() if request.method == "GET" else form_from_request()
    template = None
    students = []
    result = None
    messages = get_flashed_messages(with_categories=True)
    error = next((message for category, message in messages if category == "error"), None)
    success_message = next((message for category, message in messages if category == "success"), None)

    if request.method == "POST":
        try:
            if not form["students_folder_id"]:
                raise ValueError("Select the students folder from Drive before correcting.")
            service = get_drive_service()
            template, students = load_drive_files(service, form)
            result = grade_drive_files(service, form, template, students)
            flash(
                f"Corrected {result['graded_count']} workbook(s). Output folder: {result['corrected_folder_name']}",
                "success",
            )
            return redirect(url_for("index"))
        except Exception as exc:
            flash(str(exc), "error")
            return redirect(url_for("index"))

    return render_template_string(
        PAGE,
        error=error,
        success_message=success_message,
        corrected_folder_placeholder=f"{DEFAULT_CORRECTED_FOLDER_PREFIX}-YYYYMMDD-HHMMSS",
        folder_mime_type=FOLDER_MIME_TYPE,
        form=form,
        reset_summary_on_load=request.method == "GET",
        result=result,
        signed_in=("credentials" in session or not os.environ.get("RENDER")),
        students=students,
        template=template,
        xlsx_mime_type=XLSX_MIME_TYPE,
    )


@app.after_request
def add_no_store_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.route("/login")
def login():
    flow = make_oauth_flow()
    authorization_url, state = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
    )
    session["state"] = state
    return redirect(authorization_url)


@app.route("/oauth2callback")
def oauth2callback():
    flow = make_oauth_flow(session.get("state"))
    flow.fetch_token(authorization_response=request.url)
    credentials = flow.credentials
    session["credentials"] = credentials_to_dict(credentials)
    return redirect(url_for("index"))


@app.route("/logout")
def logout():
    session.pop("credentials", None)
    session.pop("state", None)
    return redirect(url_for("index"))


@app.route("/api/drive/children")
def api_drive_children():
    folder_id = request.args.get("folder_id", "root").strip() or "root"
    try:
        service = get_drive_service()
        return jsonify({"items": list_drive_browser_items(service, folder_id)})
    except Exception as exc:
        status_code = 401 if "Sign in" in str(exc) else 500
        return jsonify({"error": str(exc)}), status_code


@app.route("/api/correction/start", methods=["POST"])
def api_correction_start():
    form = form_from_request()
    if not form["students_folder_id"]:
        return jsonify({"error": "Select the students folder from Drive before correcting."}), 400
    credentials_data = current_credentials_data()
    if os.environ.get("RENDER") and not credentials_data:
        return jsonify({"error": "Sign in with Google Drive first."}), 401

    job_id = uuid.uuid4().hex
    with JOBS_LOCK:
        JOBS[job_id] = {
            "id": job_id,
            "status": "queued",
            "stage": "Queued...",
            "completed": 0,
            "total": 0,
            "submissions": [],
            "corrected_folder_id": "",
            "corrected_folder_name": "",
            "total_differences": 0,
            "error": None,
            "cancel_requested": False,
        }

    thread = threading.Thread(target=run_correction_job, args=(job_id, form, credentials_data), daemon=True)
    thread.start()
    return jsonify({"job_id": job_id})


@app.route("/api/correction/status/<job_id>")
def api_correction_status(job_id: str):
    job = job_snapshot(job_id)
    if not job:
        return jsonify({"error": "Correction job not found."}), 404
    return jsonify(job)


@app.route("/api/correction/cancel/<job_id>", methods=["POST"])
def api_correction_cancel(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return jsonify({"error": "Correction job not found."}), 404
        job["cancel_requested"] = True
        if job["status"] in {"queued", "running"}:
            job["stage"] = "Cancelling after the current file..."
    return jsonify({"status": "cancelling"})


@app.route("/health")
def health():
    return {"status": "ok"}


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)
